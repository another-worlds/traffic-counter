"""
xlsx export. One workbook per export request, with:
  - "Summary" sheet: every line × video pair, plus totals.
  - One sheet per video, with the same breakdown for that video alone.
  - When a video has per-hour segments, each hour gets its own sheet
    showing line counts for that hour only.
"""
from __future__ import annotations
import io
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import pandas as pd

from .counting import compute_counts_for_lines, COCO_VEHICLE_CLASSES
from .tracks import load_materialized_tracks, load_tracks_for_video, load_tracks_for_videos


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0C447C")


def _write_header(ws, headers, row=1):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_letter].width = min(max(width + 2, 10), 40)


def _line_rows(counts: Dict, scope: str) -> List[List]:
    """Flatten one compute_counts_for_lines() result into rows."""
    rows = []
    class_keys = list(COCO_VEHICLE_CLASSES.values())
    for ln in counts["per_line"]:
        rows.append([
            scope,
            ln["line_name"],
            ln["total"],
            ln["percent_of_video_total"],
            ln["percent_of_drawn_lines"],
            ln["by_direction"].get("positive", 0),
            ln["by_direction"].get("negative", 0),
            *[ln["by_class"].get(k, 0) for k in class_keys],
        ])
    return rows


def build_xlsx(
    project_id: str,
    project_name: str,
    video_rows: List[Dict],   # [{"id": str, "filename": str}, ...]
    lines: List[Dict],        # [{"id": str, "name": str, "a":[x,y], "b":[x,y]}]
) -> bytes:
    """Generate and return the xlsx workbook bytes."""
    wb = Workbook()

    # Sheet 1: Summary (aggregated across all selected videos)
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"Project: {project_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Videos: {len(video_rows)}, lines: {len(lines)}")

    class_keys = list(COCO_VEHICLE_CLASSES.values())
    headers = [
        "Scope", "Line", "Total tracks",
        "% of total in scope", "% of drawn lines (in scope)",
        "Dir +", "Dir -",
        *class_keys,
    ]
    _write_header(ws, headers, row=4)

    # Aggregated across all selected videos
    video_ids = [v["id"] for v in video_rows]
    all_tracks = load_tracks_for_videos(project_id, video_ids)
    aggregated = compute_counts_for_lines(all_tracks, lines)

    r = 5
    for row in _line_rows(aggregated, scope="ALL VIDEOS"):
        for j, val in enumerate(row, start=1):
            ws.cell(row=r, column=j, value=val)
        r += 1

    # Totals row
    ws.cell(row=r, column=1, value="").font = Font(bold=True)
    ws.cell(row=r, column=2, value="TOTAL (unique tracks across lines)").font = Font(bold=True)
    ws.cell(row=r, column=3, value=aggregated["sum_across_lines"]).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=2, value="UNIQUE TRACKS IN SCOPE").font = Font(bold=True)
    ws.cell(row=r, column=3, value=aggregated["total_unique_tracks"]).font = Font(bold=True)
    _autosize(ws)

    # One sheet per video
    for v in video_rows:
        # Use the cached materialised form: the /counts endpoint has very
        # likely already paid the cold-build cost for the video the user is
        # exporting, so each per-video sheet is effectively free.
        mt = load_materialized_tracks(project_id, v["id"])
        # In a per-video sheet, "scope" is the single video → namespacing not needed
        per = compute_counts_for_lines(mt, lines)

        title = (v["filename"][:28] + "…") if len(v["filename"]) > 28 else v["filename"]
        # openpyxl sheet titles must be unique and <= 31 chars, no /\?*[]:
        safe = "".join(c for c in title if c not in r'/\?*[]:').strip() or v["id"][:8]
        # Avoid collisions
        base = safe
        i = 1
        while base in wb.sheetnames:
            i += 1
            base = f"{safe[:28]}_{i}"
        wsv = wb.create_sheet(title=base)

        wsv.cell(row=1, column=1, value=v["filename"]).font = Font(bold=True, size=12)
        _write_header(wsv, headers, row=3)
        rr = 4
        for row in _line_rows(per, scope=v["filename"]):
            for j, val in enumerate(row, start=1):
                wsv.cell(row=rr, column=j, value=val)
            rr += 1
        wsv.cell(row=rr, column=2, value="UNIQUE TRACKS IN VIDEO").font = Font(bold=True)
        wsv.cell(row=rr, column=3, value=per["total_unique_tracks"]).font = Font(bold=True)
        _autosize(wsv)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_xlsx_for_video(
    project_name: str,
    video_filename: str,
    tracks_df,
    lines: List[Dict],
    segments: Optional[List[Dict]] = None,
) -> bytes:
    """One-video workbook: Summary header + a single details sheet.
    When ``segments`` is provided (list of dicts with start_time_s, end_time_s,
    segment_idx), one additional sheet per hour is appended showing line
    counts for that hour only.

    Used by the per-video export job.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"Project: {project_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Video: {video_filename}")
    ws.cell(row=3, column=1, value=f"Lines: {len(lines)}")

    class_keys = list(COCO_VEHICLE_CLASSES.values())
    headers = [
        "Scope", "Line", "Total tracks",
        "% of total in scope", "% of drawn lines (in scope)",
        "Dir +", "Dir -",
        *class_keys,
    ]
    _write_header(ws, headers, row=5)

    counts = compute_counts_for_lines(tracks_df, lines)
    r = 6
    for row in _line_rows(counts, scope=video_filename):
        for j, val in enumerate(row, start=1):
            ws.cell(row=r, column=j, value=val)
        r += 1

    ws.cell(row=r, column=2, value="TOTAL (unique tracks across lines)").font = Font(bold=True)
    ws.cell(row=r, column=3, value=counts["sum_across_lines"]).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=2, value="UNIQUE TRACKS IN VIDEO").font = Font(bold=True)
    ws.cell(row=r, column=3, value=counts["total_unique_tracks"]).font = Font(bold=True)
    _autosize(ws)

    # Per-hour sheets when segment metadata is available.
    if segments:
        _add_hourly_sheets(wb, tracks_df, lines, headers, segments)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _add_hourly_sheets(
    wb: Workbook,
    tracks_df,
    lines: List[Dict],
    headers: List[str],
    segments: List[Dict],
) -> None:
    """Append one sheet per segment to *wb*, filtering *tracks_df* by time range."""
    import pandas as _pd
    from .counting import materialize_tracks as _mat

    is_df = isinstance(tracks_df, _pd.DataFrame)

    for seg in sorted(segments, key=lambda s: s["segment_idx"]):
        idx = seg["segment_idx"]
        t0 = float(seg.get("start_time_s", 0))
        t1 = float(seg.get("end_time_s", t0 + 3600))

        # Format label: "Hour 0 (0:00–1:00)"
        def _fmt(secs: float) -> str:
            h = int(secs // 3600)
            m = int((secs % 3600) // 60)
            return f"{h}:{m:02d}"

        label = f"Hour {idx} ({_fmt(t0)}–{_fmt(t1)})"
        safe = "".join(c for c in label if c not in r'/\?*[]:').strip()
        base = safe[:31]
        i = 1
        while base in wb.sheetnames:
            i += 1
            base = f"{safe[:28]}_{i}"

        wsv = wb.create_sheet(title=base)
        wsv.cell(row=1, column=1, value=label).font = Font(bold=True, size=12)
        _write_header(wsv, headers, row=3)

        # Slice the full DataFrame to this hour. Prefer the exact frame_idx
        # range from the segment (integers — no float-boundary ambiguity);
        # fall back to the absolute t_seconds window for old metadata.
        if is_df and not tracks_df.empty:
            if "start_frame" in seg and "end_frame" in seg:
                f0 = int(seg["start_frame"])
                f1 = int(seg["end_frame"])
                mask = (tracks_df["frame_idx"] >= f0) & (tracks_df["frame_idx"] < f1)
            else:
                mask = (tracks_df["t_seconds"] >= t0) & (tracks_df["t_seconds"] < t1)
            seg_df = tracks_df[mask]
        else:
            seg_df = _pd.DataFrame(columns=tracks_df.columns if is_df else [
                "frame_idx", "t_seconds", "track_id", "class_id", "conf",
                "cx", "cy", "w", "h",
            ])

        seg_mt = _mat(seg_df)
        per = compute_counts_for_lines(seg_mt, lines)

        rr = 4
        for row in _line_rows(per, scope=label):
            for j, val in enumerate(row, start=1):
                wsv.cell(row=rr, column=j, value=val)
            rr += 1
        wsv.cell(row=rr, column=2, value="UNIQUE TRACKS IN HOUR").font = Font(bold=True)
        wsv.cell(row=rr, column=3, value=per["total_unique_tracks"]).font = Font(bold=True)
        _autosize(wsv)
